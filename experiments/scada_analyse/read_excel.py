from openpyxl import load_workbook
import sys

try:
    wb = load_workbook('config_client.xlsx', data_only=True)

    for sheet_name in wb.sheetnames:
        print(f'=== Sheet: {sheet_name} ===')
        ws = wb[sheet_name]

        for row in ws.iter_rows(values_only=True):
            print('\t'.join([str(cell) if cell is not None else '' for cell in row]))
        print()
except Exception as e:
    print(f'Error: {e}', file=sys.stderr)
    import traceback
    traceback.print_exc()
    sys.exit(1)
